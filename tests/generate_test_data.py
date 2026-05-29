"""Generate realistic BOM test data for benchmarking.

Creates Excel files that mimic real manufacturing BOMs with various
edge cases: merged cells, missing fields, mixed formats, different column names.
"""

import datetime
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TEST_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(TEST_DIR, "fixtures")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _style_header(ws, row, col_count):
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_data(ws, row, col_count):
    font = Font(name="宋体", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = border


# ── Test Case 1: Standard BOM (变速箱总成) ──

def generate_standard_bom():
    """Standard BOM with 25 items — typical gearbox assembly."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM清单"

    ws.merge_cells("A1:K1")
    ws.cell(row=1, column=1, value="变速箱总成 BOM清单").font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws.cell(row=2, column=1, value="设计：张三 | 日期：2026-05-20 | 版本：V2.1").font = Font(size=9, color="666666")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers = ["序号", "图号/零件号", "零件名称", "材料", "规格", "数量", "单位", "单重(kg)", "总重(kg)", "来源", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(headers))

    items = [
        [1, "BSJ-01-001", "箱体", "HT250", "400×300×200", 1, "件", 45.0, 45.0, "铸造件", "时效处理"],
        [2, "BSJ-01-002", "箱盖", "HT250", "400×300×25", 1, "件", 8.5, 8.5, "铸造件", ""],
        [3, "BSJ-01-003", "输入轴", "40Cr", "φ55×320", 1, "件", 5.8, 5.8, "自制件", "调质HRC28-32"],
        [4, "BSJ-01-004", "输出轴", "40Cr", "φ65×380", 1, "件", 8.2, 8.2, "自制件", "调质HRC28-32"],
        [5, "BSJ-01-005", "中间轴", "45#", "φ45×260", 1, "件", 3.2, 3.2, "自制件", ""],
        [6, "BSJ-01-006", "一级主动齿轮", "20CrMnTi", "m=3 z=28", 1, "件", 1.5, 1.5, "自制件", "渗碳淬火HRC58-62"],
        [7, "BSJ-01-007", "一级从动齿轮", "20CrMnTi", "m=3 z=56", 1, "件", 4.2, 4.2, "自制件", "渗碳淬火HRC58-62"],
        [8, "BSJ-01-008", "二级主动齿轮", "20CrMnTi", "m=4 z=22", 1, "件", 1.8, 1.8, "自制件", "渗碳淬火HRC58-62"],
        [9, "BSJ-01-009", "二级从动齿轮", "20CrMnTi", "m=4 z=65", 1, "件", 8.5, 8.5, "自制件", "渗碳淬火HRC58-62"],
        [10, "BSJ-01-010", "输入轴承端盖", "Q235A", "φ120×15", 1, "件", 0.8, 0.8, "自制件", ""],
        [11, "BSJ-01-011", "输出轴承端盖", "Q235A", "φ140×15", 1, "件", 1.0, 1.0, "自制件", ""],
        [12, "BSJ-01-012", "中间轴承端盖", "Q235A", "φ110×12", 2, "件", 0.6, 1.2, "自制件", ""],
        [13, "BSJ-01-013", "输入油封", "丁腈橡胶", "TC55×78×10", 1, "件", 0.03, 0.03, "外购件", "NOK品牌"],
        [14, "BSJ-01-014", "输出油封", "丁腈橡胶", "TC65×90×12", 1, "件", 0.04, 0.04, "外购件", "NOK品牌"],
        [15, "BSJ-01-015", "输入轴承", "GCr15", "6311", 2, "套", 0.85, 1.7, "外购件", "SKF"],
        [16, "BSJ-01-016", "输出轴承", "GCr15", "6313", 2, "套", 1.2, 2.4, "外购件", "SKF"],
        [17, "BSJ-01-017", "中间轴承", "GCr15", "6309", 4, "套", 0.55, 2.2, "外购件", "SKF"],
        [18, "BSJ-01-018", "通气帽", "Q235A", "M20×1.5", 1, "件", 0.08, 0.08, "外购件", ""],
        [19, "BSJ-01-019", "放油螺塞", "35CrMo", "M16×1.5", 1, "件", 0.05, 0.05, "标准件", "镀锌"],
        [20, "BSJ-01-020", "油标", "组合件", "YWZ-80", 1, "套", 0.12, 0.12, "外购件", ""],
        [21, "GB/T 5783-2012", "六角头螺栓", "8.8级", "M12×40", 16, "件", 0.04, 0.64, "标准件", "镀锌"],
        [22, "GB/T 5783-2012", "六角头螺栓", "8.8级", "M10×30", 8, "件", 0.025, 0.2, "标准件", "镀锌"],
        [23, "GB/T 93-1987", "弹簧垫圈", "65Mn", "M12", 16, "件", 0.003, 0.048, "标准件", ""],
        [24, "GB/T 97.1-2002", "平垫圈", "Q235A", "M12", 16, "件", 0.005, 0.08, "标准件", ""],
        [25, "BSJ-01-021", "密封垫片", "石棉橡胶", "400×300×2", 2, "件", 0.15, 0.3, "外购件", ""],
    ]

    for r, row_data in enumerate(items, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data(ws, r, len(headers))

    # Summary row
    summary_row = 5 + len(items)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
    ws.cell(row=summary_row, column=1, value=f"共 {len(items)} 项").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=sum(it[5] for it in items))
    ws.cell(row=summary_row, column=9, value=round(sum(it[8] for it in items), 2))

    col_widths = [6, 18, 18, 12, 16, 6, 6, 10, 10, 8, 14]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    path = os.path.join(OUTPUT_DIR, "标准BOM_变速箱总成.xlsx")
    wb.save(path)
    return path, len(items)


# ── Test Case 2: BOM with merged cells (合并单元格) ──

def generate_merged_cell_bom():
    """BOM where part_no and part_name span multiple rows (common in real drawings)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "明细表"

    headers = ["序号", "代号", "名称", "材料", "数量", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    # Merged cells: part_no/name span rows for sub-assemblies
    items = [
        # Sub-assembly 1: 底座组件 (rows 2-5, merged B2:B5, C2:C5)
        [1, "DJ-01-001", "底座", "HT200", 1, "铸造件"],
        [2, "", "", "Q235A", 2, "加强筋"],
        [3, "", "", "45#", 4, "安装支座"],
        [4, "", "", "8.8级", 8, "地脚螺栓M16"],
        # Sub-assembly 2: 传动组件 (rows 6-9)
        [5, "DJ-01-002", "主动轴", "40Cr", 1, "调质"],
        [6, "", "", "20CrMnTi", 1, "齿轮"],
        [7, "", "", "GCr15", 2, "轴承6208"],
        [8, "", "", "丁腈橡胶", 1, "油封TC40×62×8"],
        # Sub-assembly 3: 罩壳组件 (rows 10-12)
        [9, "DJ-01-003", "上罩壳", "Q235A", 1, "钣金件"],
        [10, "", "", "Q235A", 1, "下罩壳"],
        [11, "", "", "65Mn", 12, "弹簧螺母M6"],
        # Single items
        [12, "DJ-01-004", "铭牌", "铝", 1, "蚀刻"],
        [13, "DJ-01-005", "防护罩", "尼龙66", 2, ""],
    ]

    for r, row_data in enumerate(items, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data(ws, r, len(headers))

    # Merge cells for sub-assemblies
    ws.merge_cells("B2:B5")
    ws.merge_cells("C2:C5")
    ws.merge_cells("B6:B9")
    ws.merge_cells("C6:C9")
    ws.merge_cells("B10:B12")
    ws.merge_cells("C10:C12")

    col_widths = [6, 14, 14, 12, 6, 16]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    path = os.path.join(OUTPUT_DIR, "合并单元格BOM_底座组件.xlsx")
    wb.save(path)
    return path, len(items)


# ── Test Case 3: Non-standard column names (非标准列名) ──

def generate_nonstandard_columns_bom():
    """BOM with unusual column naming conventions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "物料清单"

    headers = ["编号", "物料编码", "品名", "牌号", "型号规格", "用量", "计量单位", "备注说明"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    items = [
        [1, "WL-2026-001", "液压缸", "", "HOB-63/35-200", 2, "台", "力士乐"],
        [2, "WL-2026-002", "电磁换向阀", "", "4WE6E6X/EG24N9K4", 1, "台", "力士乐"],
        [3, "WL-2026-003", "溢流阀", "", "DBDH10P10B/100", 1, "台", ""],
        [4, "WL-2026-004", "油泵", "", "PV2R2-26-F-RAA-4222", 1, "台", "不二越"],
        [5, "WL-2026-005", "电机", "", "Y132M-4-7.5kW", 1, "台", ""],
        [6, "WL-2026-006", "油箱", "Q235A", "500×400×300", 1, "件", "自制件"],
        [7, "WL-2026-007", "油管", "20#", "φ18×2 L=500", 4, "根", ""],
        [8, "WL-2026-008", "管接头", "45#", "M18×1.5-卡套式", 8, "件", ""],
        [9, "WL-2026-009", "过滤器", "", "TF-40×100", 1, "套", "回油过滤"],
        [10, "WL-2026-010", "压力表", "", "YN-60 0-16MPa", 2, "块", ""],
    ]

    for r, row_data in enumerate(items, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data(ws, r, len(headers))

    col_widths = [6, 16, 14, 10, 28, 6, 8, 14]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    path = os.path.join(OUTPUT_DIR, "非标准列名BOM_液压系统.xlsx")
    wb.save(path)
    return path, len(items)


# ── Test Case 4: Large BOM (大表50+项) ──

def generate_large_bom():
    """Large BOM with 50+ items — stress test."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = ["序号", "图号", "名称", "材料", "规格", "数量", "单位", "来源"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    materials = ["45#", "Q235A", "HT200", "40Cr", "20CrMnTi", "GCr15", "65Mn", "8.8级", "丁腈橡胶", "6061-T6"]
    sources = ["自制件", "外购件", "标准件"]
    names = [
        "底座", "上盖", "侧板", "加强筋", "安装板", "支撑架", "导向柱", "定位销",
        "轴", "齿轮", "轴承座", "端盖", "法兰", "联轴器", "皮带轮", "链轮",
        "螺栓", "螺母", "垫圈", "弹簧", "密封圈", "油封", "O型圈",
        "手柄", "旋钮", "标牌", "防护罩", "导轨", "滑块", "气缸",
    ]

    items = []
    for i in range(1, 56):
        name = names[(i - 1) % len(names)]
        mat = random.choice(materials)
        src = "标准件" if i > 45 else random.choice(["自制件", "外购件"])
        qty = random.randint(1, 8)
        spec = f"φ{random.randint(10,100)}×{random.randint(20,300)}"
        items.append([i, f"ZJ-{i:03d}", name, mat, spec, qty, "件", src])

    for r, row_data in enumerate(items, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data(ws, r, len(headers))

    col_widths = [6, 10, 12, 12, 14, 6, 6, 8]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    path = os.path.join(OUTPUT_DIR, "大型BOM_装配体50项.xlsx")
    wb.save(path)
    return path, len(items)


# ── Run all generators ──

if __name__ == "__main__":
    results = []
    for func in [generate_standard_bom, generate_merged_cell_bom, generate_nonstandard_columns_bom, generate_large_bom]:
        path, count = func()
        results.append((os.path.basename(path), count, path))
        print(f"  Generated: {os.path.basename(path)} ({count} items)")

    print(f"\nAll fixtures saved to: {OUTPUT_DIR}")
